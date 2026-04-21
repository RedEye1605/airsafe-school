#!/usr/bin/env python3
"""
Update AirSafe School Progress Tracker for Day 1 completion.
"""
import openpyxl

# File path (adjust if needed)
TRACKER_PATH = "/tmp/datathon-elevate/Datathon Elevate/Airsafe Team Progress Tracker.xlsm"

# Load the workbook (read_only=True to preserve formulas/macros)
wb = openpyxl.load_workbook(TRACKER_PATH, read_only=True, keep_vba=True)

# Find Rhendy - Data sheet
sheet_name = 'Rhendy - Data'
if sheet_name not in wb.sheetnames:
    print(f"ERROR: Sheet '{sheet_name}' not found")
    print(f"Available sheets: {wb.sheetnames}")
    exit(1)

ws = wb[sheet_name]

# Define updates: (Task ID row number from earlier output, New Status, New % Complete)
# Looking at the earlier read output, tasks start at row 11
updates = [
    {
        'task_id': 'T-D1-03',
        'status': 'Done',
        'percent': 1.0,
        'evidence': 'spku_sample.json + GeoJSON files',
        'next_action': 'Move to Day 2 tasks'
    },
    {
        'task_id': 'T-D1-04',
        'status': 'Done',
        'percent': 1.0,
        'evidence': '3 CSV files: sd_dki.csv (2,163), smp_dki.csv (778), sma_smk_dki.csv (1,033)',
        'next_action': 'Start Day 2 geocoding'
    }
]

# Column indices (based on earlier read output)
# Columns: A=TaskID, B=Phase/Bucket, C=DueDate, D=Task/Item, E=Deliverable,
# F=Priority, G=Status, H=%Complete, I=NextAction, J=Evidence/Link, K=Notes
STATUS_COL = 7  # Column G
PERCENT_COL = 8  # Column H
NEXT_ACTION_COL = 9  # Column I
EVIDENCE_COL = 11  # Column K (note: J might be used differently)

# Create a new workbook for writing (can't write to .xlsm with macros)
wb_write = openpyxl.Workbook()
ws_write = wb_write.active

# Copy all data to new sheet
for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    ws_write.append(row)

# Apply updates
updated_tasks = []
for update in updates:
    # Find the row by Task ID (column A)
    for row_idx, row in enumerate(ws_write.iter_rows(min_row=2, values_only=True)):
        if row[0] == update['task_id']:
            # Update status, percent, next_action, evidence
            # Note: Excel rows are 1-indexed in openpyxl, but data starts at row 2
            actual_row_idx = row_idx + 2

            # Update status (column G = 7)
            ws_write.cell(row=actual_row_idx, column=STATUS_COL + 1, value=update['status'])

            # Update percent (column H = 8)
            ws_write.cell(row=actual_row_idx, column=PERCENT_COL + 1, value=update['percent'])

            # Update next action (column I = 9)
            ws_write.cell(row=actual_row_idx, column=NEXT_ACTION_COL + 1, value=update['next_action'])

            # Update evidence (column K = 11, if column J exists, skip J)
            try:
                ws_write.cell(row=actual_row_idx, column=EVIDENCE_COL + 1, value=update['evidence'])
            except:
                pass

            updated_tasks.append(update['task_id'])
            print(f"Updated: {update['task_id']} -> {update['status']}")
            break

# Save as .xlsx (can't save .xlsm with VBA)
output_path = TRACKER_PATH.replace('.xlsm', '_updated.xlsx')
wb_write.save(output_path)

print(f"\nTracker updated!")
print(f"Output saved to: {output_path}")
print(f"Updated tasks: {updated_tasks}")
print(f"\nNote: Original .xlsm file preserved (VBA macros cannot be edited)")
print(f"Please review the updated file and then manually update your .xlsm if needed")
