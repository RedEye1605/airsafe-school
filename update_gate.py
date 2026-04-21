#!/usr/bin/env python3
"""
Update Day 1 Gate status in Shared & Gates sheet.
"""
import openpyxl

TRACKER_PATH = "/tmp/datathon-elevate/Datathon Elevate/Airsafe Team Progress Tracker_updated.xlsx"

wb = openpyxl.load_workbook(TRACKER_PATH)
sheet_name = 'Shared & Gates'

if sheet_name not in wb.sheetnames:
    print(f"ERROR: Sheet '{sheet_name}' not found")
    exit(1)

ws = wb[sheet_name]

# Find M-D1-GATE row (Day 1 Gate)
# Looking for task ID "M-D1-GATE"
for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row[0] == 'M-D1-GATE':
        actual_row_idx = row_idx + 2
        # Update status (column G = 7) to "Done"
        ws.cell(row=actual_row_idx, column=8, value='Done')
        # Update % complete (column H = 8) to 100%
        ws.cell(row=actual_row_idx, column=9, value=1.0)
        print(f"Updated: M-D1-GATE -> Done (100%)")
        break

# Save
wb.save(TRACKER_PATH)
print(f"Gate status updated in {TRACKER_PATH}")
